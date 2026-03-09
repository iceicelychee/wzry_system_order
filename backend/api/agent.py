import io
import uuid
import random
import string
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from models import (
    get_db, Admin, Agent, Order, BalanceLog, OperationLog,
    AgentStatus, LinkStatus, ExecStatus,
)
from services.auth import (
    get_password_hash, verify_password, create_access_token,
    get_current_admin, get_current_agent,
)
from services.order_utils import generate_order_no
from config import settings

router = APIRouter(prefix="/api/agent", tags=["代理管理"])


# ============== Pydantic Models ==============

class CreateAgentRequest(BaseModel):
    username: str
    password: str
    balance: int = 0
    remark: Optional[str] = None


class UpdateAgentRequest(BaseModel):
    status: Optional[str] = None
    remark: Optional[str] = None


class RechargeRequest(BaseModel):
    amount: int
    reason: Optional[str] = None


class ResetPasswordRequest(BaseModel):
    new_password: str


class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


class BatchCreateOrderRequest(BaseModel):
    count: int
    remark: Optional[str] = None


# ============== Helper Functions ==============

def agent_to_dict(agent: Agent) -> dict:
    return {
        "id": agent.id,
        "username": agent.username,
        "balance": agent.balance,
        "status": agent.status.value if agent.status else "启用",
        "remark": agent.remark,
        "created_at": agent.created_at.strftime("%Y-%m-%d %H:%M:%S") if agent.created_at else None,
    }


def order_to_dict(order: Order) -> dict:
    return {
        "id": order.id,
        "order_no": order.order_no,
        "token": order.token,
        "client_link": f"{settings.FRONTEND_URL}/order/{order.token}",
        "system_type": order.system_type.value if order.system_type else None,
        "image_path": order.image_path,
        "image_url": f"/uploads/orders/{order.image_path}" if order.image_path else None,
        "link_status": order.link_status.value if order.link_status else "未填写",
        "exec_status": order.exec_status.value if order.exec_status else "待执行",
        "remark": order.remark,
        "created_at": order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else None,
        "submitted_at": order.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if order.submitted_at else None,
        "agent_id": order.agent_id,
        "agent_name": order.agent.username if order.agent else None,
    }


# ============== 管理员操作代理的接口 ==============

@router.post("/create")
def create_agent(
    req: CreateAgentRequest,
    current_admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    # 检查用户名是否已存在
    if db.query(Agent).filter(Agent.username == req.username).first():
        raise HTTPException(status_code=400, detail="用户名已存在")
    
    agent = Agent(
        username=req.username,
        password=get_password_hash(req.password),
        balance=req.balance,
        remark=req.remark,
        status=AgentStatus.enabled,
    )
    db.add(agent)
    db.commit()
    db.refresh(agent)
    
    # 记录初始余额
    if req.balance > 0:
        log = BalanceLog(
            agent_id=agent.id,
            change_amount=req.balance,
            balance_after=req.balance,
            reason="初始余额",
            operator=current_admin.username,
        )
        db.add(log)
        db.commit()
    
    db.add(OperationLog(admin_id=current_admin.id, action="创建代理", target=req.username))
    db.commit()
    
    return {"message": "代理创建成功", "agent": agent_to_dict(agent)}


@router.get("/list")
def list_agents(
    page: int = 1,
    page_size: int = 20,
    keyword: str = "",
    status: str = "",
    current_admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    query = db.query(Agent)
    
    if keyword:
        query = query.filter(Agent.username.contains(keyword))
    if status:
        query = query.filter(Agent.status == status)
    
    total = query.count()
    agents = query.order_by(desc(Agent.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "list": [agent_to_dict(a) for a in agents],
    }


@router.get("/detail/{agent_id}")
def get_agent_detail(
    agent_id: int,
    current_admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    agent = db.query(Agent).filter(Agent.id == agent_id).with_for_update().first()
    if not agent:
        raise HTTPException(status_code=404, detail="代理不存在")
    
    # 获取最近余额变动记录
    logs = db.query(BalanceLog).filter(BalanceLog.agent_id == agent_id).order_by(desc(BalanceLog.created_at)).limit(20).all()
    
    return {
        "agent": agent_to_dict(agent),
        "balance_logs": [
            {
                "id": log.id,
                "change_amount": log.change_amount,
                "balance_after": log.balance_after,
                "reason": log.reason,
                "operator": log.operator,
                "order_id": log.order_id,
                "created_at": log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else None,
            }
            for log in logs
        ],
    }


@router.put("/update/{agent_id}")
def update_agent(
    agent_id: int,
    req: UpdateAgentRequest,
    current_admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    agent = db.query(Agent).filter(Agent.id == agent_id).with_for_update().first()
    if not agent:
        raise HTTPException(status_code=404, detail="代理不存在")
    
    if req.status:
        agent.status = AgentStatus(req.status)
    if req.remark is not None:
        agent.remark = req.remark
    
    db.commit()
    db.add(OperationLog(admin_id=current_admin.id, action="修改代理", target=agent.username))
    db.commit()
    
    return {"message": "修改成功"}


@router.post("/recharge/{agent_id}")
def recharge_agent(
    agent_id: int,
    req: RechargeRequest,
    current_admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    agent = db.query(Agent).filter(Agent.id == agent_id).with_for_update().first()
    if not agent:
        raise HTTPException(status_code=404, detail="代理不存在")
    
    new_balance = agent.balance + req.amount
    if new_balance < 0:
        raise HTTPException(status_code=400, detail="扣减后余额不能为负")
    
    agent.balance = new_balance
    reason = req.reason or ("管理员充值" if req.amount > 0 else "管理员扣减")
    
    log = BalanceLog(
        agent_id=agent.id,
        change_amount=req.amount,
        balance_after=new_balance,
        reason=reason,
        operator=current_admin.username,
    )
    db.add(log)
    db.commit()
    
    db.add(OperationLog(
        admin_id=current_admin.id,
        action="调整代理余额",
        target=f"{agent.username}: {req.amount:+d}",
    ))
    db.commit()
    
    return {"message": "余额调整成功", "balance": new_balance}


@router.delete("/delete/{agent_id}")
def delete_agent(
    agent_id: int,
    current_admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    agent = db.query(Agent).filter(Agent.id == agent_id).with_for_update().first()
    if not agent:
        raise HTTPException(status_code=404, detail="代理不存在")
    
    # 检查是否有关联订单
    order_count = db.query(Order).filter(Order.agent_id == agent_id).count()
    if order_count > 0:
        raise HTTPException(status_code=400, detail=f"该代理有 {order_count} 个关联订单，无法删除")
    
    username = agent.username
    db.delete(agent)
    db.add(OperationLog(admin_id=current_admin.id, action="删除代理", target=username))
    db.commit()
    
    return {"message": "代理已删除"}


@router.post("/reset-password/{agent_id}")
def reset_agent_password(
    agent_id: int,
    req: ResetPasswordRequest,
    current_admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    agent = db.query(Agent).filter(Agent.id == agent_id).with_for_update().first()
    if not agent:
        raise HTTPException(status_code=404, detail="代理不存在")
    
    agent.password = get_password_hash(req.new_password)
    db.commit()
    
    db.add(OperationLog(admin_id=current_admin.id, action="重置代理密码", target=agent.username))
    db.commit()
    
    return {"message": "密码重置成功"}


# ============== 代理自身操作的接口 ==============

@router.get("/me")
def get_agent_me(current_agent: Agent = Depends(get_current_agent)):
    return {
        "id": current_agent.id,
        "username": current_agent.username,
        "balance": current_agent.balance,
        "status": current_agent.status.value if current_agent.status else "启用",
        "created_at": current_agent.created_at.strftime("%Y-%m-%d %H:%M:%S") if current_agent.created_at else None,
    }


@router.post("/change-password")
def agent_change_password(
    req: ChangePasswordRequest,
    current_agent: Agent = Depends(get_current_agent),
    db: Session = Depends(get_db),
):
    if not verify_password(req.old_password, current_agent.password):
        raise HTTPException(status_code=400, detail="原密码错误")
    current_agent.password = get_password_hash(req.new_password)
    db.commit()
    return {"message": "密码修改成功"}


@router.get("/my-balance-logs")
def get_my_balance_logs(
    page: int = 1,
    page_size: int = 20,
    current_agent: Agent = Depends(get_current_agent),
    db: Session = Depends(get_db),
):
    query = db.query(BalanceLog).filter(BalanceLog.agent_id == current_agent.id)
    total = query.count()
    logs = query.order_by(desc(BalanceLog.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "list": [
            {
                "id": log.id,
                "change_amount": log.change_amount,
                "balance_after": log.balance_after,
                "reason": log.reason,
                "operator": log.operator,
                "order_id": log.order_id,
                "created_at": log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else None,
            }
            for log in logs
        ],
    }


@router.get("/my-orders")
def get_my_orders(
    page: int = 1,
    page_size: int = 20,
    link_status: str = "",
    exec_status: str = "",
    keyword: str = "",
    current_agent: Agent = Depends(get_current_agent),
    db: Session = Depends(get_db),
):
    query = db.query(Order).filter(Order.agent_id == current_agent.id)
    
    if link_status:
        query = query.filter(Order.link_status == link_status)
    if exec_status:
        query = query.filter(Order.exec_status == exec_status)
    if keyword:
        query = query.filter(Order.order_no.contains(keyword))
    
    total = query.count()
    orders = query.order_by(desc(Order.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "list": [order_to_dict(o) for o in orders],
    }


@router.get("/my-orders/{order_id}")
def get_my_order_detail(
    order_id: int,
    current_agent: Agent = Depends(get_current_agent),
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id, Order.agent_id == current_agent.id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    
    # 获取执行日志
    logs = db.query(ExecLog).filter(ExecLog.order_id == order.id).order_by(ExecLog.created_at).limit(50).all()
    
    # 获取回调记录
    callbacks = db.query(ScriptCallback).filter(ScriptCallback.order_id == order.id).order_by(desc(ScriptCallback.created_at)).all()
    
    return {
        **order_to_dict(order),
        "exec_logs": [
            {"content": log.log_content, "time": log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else None}
            for log in logs
        ],
        "callbacks": [
            {
                "id": cb.id,
                "qrcode_status": cb.qrcode_status.value if cb.qrcode_status else None,
                "qrcode_url": f"/uploads/qrcodes/{cb.qrcode_path}" if cb.qrcode_path else None,
                "qrcode_expired_at": cb.qrcode_expired_at.strftime("%Y-%m-%d %H:%M:%S") if cb.qrcode_expired_at else None,
                "result": cb.result,
                "error_msg": cb.error_msg,
                "created_at": cb.created_at.strftime("%Y-%m-%d %H:%M:%S") if cb.created_at else None,
            }
            for cb in callbacks
        ],
    }


@router.post("/my-orders/create")
async def create_my_order(
    order_no: Optional[str] = Form(None),
    system_type: Optional[str] = Form(None),
    remark: Optional[str] = Form(None),
    image: Optional[UploadFile] = File(None),
    current_agent: Agent = Depends(get_current_agent),
    db: Session = Depends(get_db),
):
    # 检查余额
    if current_agent.balance < 1:
        raise HTTPException(status_code=400, detail="余额不足，请充值后继续")
    
    # 使用行锁防止并发
    agent = db.query(Agent).filter(Agent.id == current_agent.id).with_for_update().first()
    if agent.balance < 1:
        raise HTTPException(status_code=400, detail="余额不足，请充值后继续")
    
    # 生成订单编号
    if order_no:
        if db.query(Order).filter(Order.order_no == order_no).first():
            raise HTTPException(status_code=400, detail="订单编号已存在")
    else:
        order_no = generate_order_no()
    
    # 处理图片上传
    image_path = None
    if image:
        ext = image.filename.rsplit(".", 1)[-1].lower()
        if ext not in ["jpg", "jpeg", "png", "bmp", "gif", "webp"]:
            raise HTTPException(status_code=400, detail="不支持的图片格式")
        image_path = f"{uuid.uuid4().hex}.{ext}"
        import aiofiles
        import os
        os.makedirs(os.path.join(settings.UPLOAD_DIR, "orders"), exist_ok=True)
        async with aiofiles.open(os.path.join(settings.UPLOAD_DIR, "orders", image_path), "wb") as f:
            await f.write(await image.read())
    
    # 创建订单
    token = uuid.uuid4().hex
    order = Order(
        order_no=order_no,
        token=token,
        system_type=system_type,
        image_path=image_path,
        remark=remark,
        agent_id=current_agent.id,
        link_status=LinkStatus.submitted if system_type and image_path else LinkStatus.unfilled,
        exec_status=ExecStatus.pending,
    )
    db.add(order)
    
    # 扣减余额
    agent.balance -= 1
    
    # 记录余额变动
    balance_log = BalanceLog(
        agent_id=agent.id,
        change_amount=-1,
        balance_after=agent.balance,
        reason="创建订单扣费",
        operator="system",
    )
    db.add(balance_log)
    
    db.commit()
    db.refresh(order)
    
    return {"message": "订单创建成功", "order": order_to_dict(order)}


@router.post("/my-orders/batch-create")
def batch_create_my_orders(
    req: BatchCreateOrderRequest,
    current_agent: Agent = Depends(get_current_agent),
    db: Session = Depends(get_db),
):
    if req.count < 1 or req.count > 100:
        raise HTTPException(status_code=400, detail="创建数量需在1-100之间")
    
    # 使用行锁防止并发
    agent = db.query(Agent).filter(Agent.id == current_agent.id).with_for_update().first()
    
    if agent.balance < req.count:
        raise HTTPException(status_code=400, detail=f"余额不足，当前余额 {agent.balance} 点，需要 {req.count} 点")
    
    orders = []
    for _ in range(req.count):
        order = Order(
            order_no=generate_order_no(),
            token=uuid.uuid4().hex,
            remark=req.remark,
            agent_id=current_agent.id,
            link_status=LinkStatus.unfilled,
            exec_status=ExecStatus.pending,
        )
        db.add(order)
        orders.append(order)
    
    # 扣减余额
    agent.balance -= req.count
    
    # 记录余额变动
    balance_log = BalanceLog(
        agent_id=agent.id,
        change_amount=-req.count,
        balance_after=agent.balance,
        reason=f"批量创建{req.count}个订单",
        operator="system",
    )
    db.add(balance_log)
    
    db.commit()
    
    return {"message": f"成功创建{req.count}个订单", "count": req.count, "balance": agent.balance}


@router.get("/my-orders/export")
def export_my_orders(
    link_status: str = "",
    exec_status: str = "",
    keyword: str = "",
    current_agent: Agent = Depends(get_current_agent),
    db: Session = Depends(get_db),
):
    query = db.query(Order).filter(Order.agent_id == current_agent.id)
    
    if link_status:
        query = query.filter(Order.link_status == link_status)
    if exec_status:
        query = query.filter(Order.exec_status == exec_status)
    if keyword:
        query = query.filter(Order.order_no.contains(keyword))
    
    orders = query.order_by(desc(Order.created_at)).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "订单列表"
    
    headers = ["序号", "订单编号", "专属链接", "链接状态", "系统类型", "执行状态", "备注", "创建时间", "提交时间"]
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for idx, order in enumerate(orders, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=order.order_no)
        ws.cell(row=idx + 1, column=3, value=f"{settings.FRONTEND_URL}/order/{order.token}")
        ws.cell(row=idx + 1, column=4, value=order.link_status.value if order.link_status else "")
        ws.cell(row=idx + 1, column=5, value=order.system_type.value if order.system_type else "")
        ws.cell(row=idx + 1, column=6, value=order.exec_status.value if order.exec_status else "")
        ws.cell(row=idx + 1, column=7, value=order.remark or "")
        ws.cell(row=idx + 1, column=8, value=order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else "")
        ws.cell(row=idx + 1, column=9, value=order.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if order.submitted_at else "")
    
    col_widths = [8, 22, 50, 10, 10, 10, 20, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"my_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
