import uuid
from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc
from pydantic import BaseModel
import aiofiles
import os
import io

from models import get_db, Order, ExecLog, ScriptCallback, OperationLog, SystemType, LinkStatus, ExecStatus, Agent
from services.auth import get_current_admin
from services.order_utils import generate_order_no
from config import settings
from sqlalchemy.orm import joinedload

router = APIRouter(prefix="/api/order", tags=["订单管理"])

ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"}


class OrderCreateRequest(BaseModel):
    order_no: Optional[str] = None
    system_type: Optional[SystemType] = None
    remark: Optional[str] = None


class OrderUpdateRequest(BaseModel):
    system_type: Optional[SystemType] = None
    remark: Optional[str] = None
    link_status: Optional[LinkStatus] = None
    exec_status: Optional[ExecStatus] = None


def order_to_dict(order: Order, base_url: str = "") -> dict:
    latest_cb = order.callbacks[-1] if order.callbacks else None
    return {
        "id": order.id,
        "order_no": order.order_no,
        "client_link": f"{settings.FRONTEND_URL}/order/{order.token}",
        "token": order.token,
        "system_type": order.system_type,
        "image_path": order.image_path,
        "image_url": f"/uploads/orders/{os.path.basename(order.image_path)}" if order.image_path else None,
        "link_status": order.link_status,
        "exec_status": order.exec_status,
        "remark": order.remark,
        "created_at": order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else None,
        "submitted_at": order.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if order.submitted_at else None,
        "qrcode_status": latest_cb.qrcode_status if latest_cb else None,
        "qrcode_url": f"/uploads/qrcodes/{os.path.basename(latest_cb.qrcode_path)}" if latest_cb and latest_cb.qrcode_path else None,
        "agent_id": order.agent_id,
        "agent_name": order.agent.username if order.agent else None,
    }


@router.post("/create")
async def create_order(
    order_no: Optional[str] = Form(None),
    system_type: Optional[SystemType] = Form(None),
    remark: Optional[str] = Form(None),
    image: Optional[UploadFile] = File(None),
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    # 生成或使用自定义订单编号
    if order_no and order_no.strip():
        final_order_no = order_no.strip()
        if db.query(Order).filter(Order.order_no == final_order_no).first():
            raise HTTPException(status_code=400, detail="订单编号已存在")
    else:
        for _ in range(10):
            final_order_no = generate_order_no()
            if not db.query(Order).filter(Order.order_no == final_order_no).first():
                break
        else:
            raise HTTPException(status_code=500, detail="订单编号生成失败，请重试")

    token = uuid.uuid4().hex
    image_path = None

    # 处理图片上传
    if image and image.filename:
        ext = os.path.splitext(image.filename)[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=400, detail="不支持的图片格式")
        file_name = f"{uuid.uuid4().hex}{ext}"
        save_path = os.path.join(settings.UPLOAD_DIR, "orders", file_name)
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        async with aiofiles.open(save_path, "wb") as f:
            await f.write(await image.read())
        image_path = save_path

    order = Order(
        order_no=final_order_no,
        token=token,
        system_type=system_type,
        image_path=image_path,
        link_status=LinkStatus.submitted if (system_type and image_path) else LinkStatus.unfilled,
        exec_status=ExecStatus.pending,
        remark=remark,
    )
    db.add(order)
    db.flush()

    # 记录操作日志
    log = OperationLog(admin_id=current_admin.id, action="创建订单", target=final_order_no)
    db.add(log)
    db.commit()
    db.refresh(order)

    return {"message": "订单创建成功", "order": order_to_dict(order)}


@router.get("/list")
def get_order_list(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    link_status: Optional[LinkStatus] = None,
    exec_status: Optional[ExecStatus] = None,
    system_type: Optional[SystemType] = None,
    keyword: Optional[str] = None,
    agent_id: Optional[int] = None,
    creator: Optional[str] = None,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    query = db.query(Order)
    if link_status:
        query = query.filter(Order.link_status == link_status)
    if exec_status:
        query = query.filter(Order.exec_status == exec_status)
    if system_type:
        query = query.filter(Order.system_type == system_type)
    if keyword:
        # 转义 SQL LIKE 特殊字符
        keyword = keyword.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        query = query.filter(Order.order_no.contains(keyword))
    if agent_id:
        query = query.filter(Order.agent_id == agent_id)
    if creator:
        if creator == "admin":
            query = query.filter(Order.agent_id == None)
        else:
            agent = db.query(Agent).filter(Agent.username == creator).first()
            if agent:
                query = query.filter(Order.agent_id == agent.id)
            else:
                query = query.filter(Order.agent_id == -1)

    # 使用 joinedload 预加载 agent 关系，避免 N+1 查询
    query = query.options(joinedload(Order.agent))
    total = query.count()
    orders = query.order_by(desc(Order.created_at)).offset((page- 1) * page_size).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "list": [order_to_dict(o) for o in orders],
    }


@router.get("/detail/{order_id}")
def get_order_detail(
    order_id: int,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")

    logs = db.query(ExecLog).filter(ExecLog.order_id == order_id).order_by(desc(ExecLog.created_at)).limit(50).all()
    result = order_to_dict(order)
    result["exec_logs"] = [
        {"content": l.log_content, "time": l.created_at.strftime("%Y-%m-%d %H:%M:%S")} for l in logs
    ]
    result["callbacks"] = [
        {
            "id": cb.id,
            "qrcode_status": cb.qrcode_status,
            "qrcode_url": f"/uploads/qrcodes/{os.path.basename(cb.qrcode_path)}" if cb.qrcode_path else None,
            "qrcode_expired_at": cb.qrcode_expired_at.strftime("%Y-%m-%d %H:%M:%S") if cb.qrcode_expired_at else None,
            "result": cb.result,
            "error_msg": cb.error_msg,
            "created_at": cb.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        }
        for cb in order.callbacks
    ]
    return result


@router.put("/update/{order_id}")
def update_order(
    order_id: int,
    req: OrderUpdateRequest,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    if req.system_type is not None:
        order.system_type = req.system_type
    if req.remark is not None:
        order.remark = req.remark
    if req.link_status is not None:
        order.link_status = req.link_status
    if req.exec_status is not None:
        order.exec_status = req.exec_status
    db.add(OperationLog(admin_id=current_admin.id, action="修改订单", target=order.order_no))
    db.commit()
    return {"message": "更新成功"}


@router.post("/disable/{order_id}")
def disable_order_link(
    order_id: int,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    order.link_status = LinkStatus.disabled
    db.add(OperationLog(admin_id=current_admin.id, action="禁用链接", target=order.order_no))
    db.commit()
    return {"message": "链接已禁用"}


@router.delete("/delete/{order_id}")
def delete_order(
    order_id: int,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    order_no = order.order_no
    # 先删除关联记录
    db.query(ExecLog).filter(ExecLog.order_id == order_id).delete()
    db.query(ScriptCallback).filter(ScriptCallback.order_id == order_id).delete()
    db.delete(order)
    db.add(OperationLog(admin_id=current_admin.id, action="删除订单", target=order_no))
    db.commit()
    return {"message": "订单已删除"}


@router.post("/retry/{order_id}")
def retry_order(
    order_id: int,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    if not order.system_type or not order.image_path:
        raise HTTPException(status_code=400, detail="订单信息不完整，无法执行")
    order.exec_status = ExecStatus.pending
    db.add(OperationLog(admin_id=current_admin.id, action="重新执行", target=order.order_no))
    db.add(ExecLog(order_id=order.id, log_content="管理员手动触发重新执行"))
    db.commit()
    return {"message": "已重新加入执行队列"}


class BatchDeleteRequest(BaseModel):
    ids: List[int]


@router.post("/batch-delete")
def batch_delete_orders(
    req: BatchDeleteRequest,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if not req.ids:
        raise HTTPException(status_code=400, detail="请选择要删除的订单")
    orders = db.query(Order).filter(Order.id.in_(req.ids)).all()
    if not orders:
        raise HTTPException(status_code=404, detail="未找到指定订单")
    count = len(orders)
    order_ids = [o.id for o in orders]
    # 先删除关联记录
    db.query(ExecLog).filter(ExecLog.order_id.in_(order_ids)).delete(synchronize_session=False)
    db.query(ScriptCallback).filter(ScriptCallback.order_id.in_(order_ids)).delete(synchronize_session=False)
    for o in orders:
        db.delete(o)
    db.add(OperationLog(admin_id=current_admin.id, action=f"批量删除{count}个订单", target=f"共{count}条"))
    db.commit()
    return {"message": f"成功删除{count}个订单", "count": count}


class BatchCreateRequest(BaseModel):
    count: int
    remark: Optional[str] = None


@router.post("/batch-create")
def batch_create_orders(
    req: BatchCreateRequest,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if req.count < 1 or req.count > 100:
        raise HTTPException(status_code=400, detail="批量创建数量必须在1-100之间")

    created = []
    for _ in range(req.count):
        for attempt in range(10):
            order_no = generate_order_no()
            if not db.query(Order).filter(Order.order_no == order_no).first():
                break
        else:
            continue

        token = uuid.uuid4().hex
        order = Order(
            order_no=order_no,
            token=token,
            link_status=LinkStatus.unfilled,
            exec_status=ExecStatus.pending,
            remark=req.remark,
        )
        db.add(order)
        db.flush()
        created.append(order)

    db.add(OperationLog(
        admin_id=current_admin.id,
        action=f"批量创建{len(created)}个订单",
        target=f"{created[0].order_no}等" if created else "",
    ))
    db.commit()

    for o in created:
        db.refresh(o)

    return {
        "message": f"成功创建{len(created)}个订单",
        "count": len(created),
        "orders": [order_to_dict(o) for o in created],
    }


@router.get("/export")
def export_orders(
    link_status: Optional[LinkStatus] = None,
    exec_status: Optional[ExecStatus] = None,
    system_type: Optional[SystemType] = None,
    keyword: Optional[str] = None,
    creator: Optional[str] = None,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    query = db.query(Order)
    if link_status:
        query = query.filter(Order.link_status == link_status)
    if exec_status:
        query = query.filter(Order.exec_status == exec_status)
    if system_type:
        query = query.filter(Order.system_type == system_type)
    if keyword:
        # 转义 SQL LIKE 特殊字符
        keyword= keyword.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        query = query.filter(Order.order_no.contains(keyword))
    if creator:
        if creator == "admin":
            query = query.filter(Order.agent_id == None)
        else:
            agent = db.query(Agent).filter(Agent.username == creator).first()
            if agent:
                query = query.filter(Order.agent_id == agent.id)
            else:
                query = query.filter(Order.agent_id == -1)

    # 分批查询避免内存溢出
    batch_size= 1000
    offset = 0
    orders = []
    while True:
        batch = query.order_by(desc(Order.created_at)).offset(offset).limit(batch_size).all()
        orders.extend(batch)
        if len(batch) < batch_size:
            break
        offset += batch_size

    wb = Workbook()
    ws = wb.active
    ws.title = "订单列表"

    # 表头
    headers = ["订单编号", "专属链接", "链接状态", "系统类型", "执行状态", "创建者", "备注", "创建时间", "提交时间"]
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.order_no)
        ws.cell(row=row, column=2, value=f"{settings.FRONTEND_URL}/order/{order.token}")
        ws.cell(row=row, column=3, value=order.link_status.value if order.link_status else "")
        ws.cell(row=row, column=4, value=order.system_type.value if order.system_type else "")
        ws.cell(row=row, column=5, value=order.exec_status.value if order.exec_status else "")
        ws.cell(row=row, column=6, value=order.agent.username if order.agent else "管理员")
        ws.cell(row=row, column=7, value=order.remark or "")
        ws.cell(row=row, column=8, value=order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else "")
        ws.cell(row=row, column=9, value=order.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if order.submitted_at else "")

    # 列宽
    col_widths = [22, 50, 10, 10, 10, 12, 20, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    db.add(OperationLog(admin_id=current_admin.id, action="导出订单", target=f"共{len(orders)}条"))
    db.commit()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


class ExportByIdsRequest(BaseModel):
    ids: List[int]


@router.post("/export-by-ids")
def export_orders_by_ids(
    req: ExportByIdsRequest,
    current_admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    if not req.ids:
        raise HTTPException(status_code=400, detail="请选择要导出的订单")

    orders = db.query(Order).filter(Order.id.in_(req.ids)).order_by(desc(Order.created_at)).all()
    if not orders:
        raise HTTPException(status_code=404, detail="未找到指定订单")

    wb = Workbook()
    ws = wb.active
    ws.title = "订单列表"

    headers = ["序号", "订单编号", "专属链接", "链接状态", "系统类型", "执行状态", "创建者", "备注", "创建时间", "提交时间"]
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, order in enumerate(orders, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=order.order_no)
        ws.cell(row=idx + 1, column=3, value=f"{settings.FRONTEND_URL}/order/{order.token}")
        ws.cell(row=idx + 1, column=4, value=order.link_status.value if order.link_status else "")
        ws.cell(row=idx + 1, column=5, value=order.system_type.value if order.system_type else "")
        ws.cell(row=idx + 1, column=6, value=order.exec_status.value if order.exec_status else "")
        ws.cell(row=idx + 1, column=7, value=order.agent.username if order.agent else "管理员")
        ws.cell(row=idx + 1, column=8, value=order.remark or "")
        ws.cell(row=idx + 1, column=9, value=order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else "")
        ws.cell(row=idx + 1, column=10, value=order.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if order.submitted_at else "")

    col_widths = [8, 22, 50, 10, 10, 10, 12, 20, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    db.add(OperationLog(admin_id=current_admin.id, action="按选择导出订单", target=f"共{len(orders)}条"))
    db.commit()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"orders_selected_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
